import csv
import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render

from .forms import ImportForm, PostForm
from .models import Post, Proposal


def _is_admin(u):
    return u.is_authenticated and (getattr(u, "is_admin_role", False) or u.is_superuser)


admin_required = user_passes_test(_is_admin, login_url="feed")


# ── Mapeo de categorías escritas en español libre ──
_CATEGORY_ALIASES = {
    "noticias": Post.Category.NOTICIAS,
    "noticia": Post.Category.NOTICIAS,
    "circular": Post.Category.CIRCULAR,
    "circulares": Post.Category.CIRCULAR,
    "evento": Post.Category.EVENTO,
    "eventos": Post.Category.EVENTO,
    "bogota eats": Post.Category.EVENTO,
    "marketing": Post.Category.MARKETING,
}


def _norm_category(value):
    if not value:
        return Post.Category.NOTICIAS
    key = str(value).strip().lower()
    return _CATEGORY_ALIASES.get(key, Post.Category.NOTICIAS)


@login_required
def feed_list(request):
    if request.method == "POST":  # propuesta de un socio
        text = request.POST.get("text", "").strip()
        if text:
            Proposal.objects.create(author=request.user, text=text)
            messages.success(request, "Propuesta enviada exitosamente")
        return redirect("feed")

    context = {
        "active": "posts",
        "posts": Post.objects.filter(is_published=True).select_related("author"),
        "is_admin": _is_admin(request.user),
    }
    if context["is_admin"]:
        context["form"] = PostForm()
        context["import_form"] = ImportForm()
    return render(request, "feed/list.html", context)


@login_required
def post_react(request, pk):
    post = get_object_or_404(Post, pk=pk, is_published=True)
    Post.objects.filter(pk=post.pk).update(reactions=post.reactions + 1)
    return redirect("feed")


@admin_required
def post_create(request):
    form = PostForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        post = form.save(commit=False)
        post.author = request.user
        post.save()
        messages.success(request, "Noticia publicada y visible para todos los socios")
        return redirect("feed")
    return render(
        request,
        "feed/post_form.html",
        {"active": "posts", "form": form, "titulo": "Nueva publicación"},
    )


@admin_required
def post_edit(request, pk):
    post = get_object_or_404(Post, pk=pk)
    form = PostForm(request.POST or None, request.FILES or None, instance=post)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Noticia actualizada")
        return redirect("feed")
    return render(
        request,
        "feed/post_form.html",
        {"active": "posts", "form": form, "titulo": "Editar publicación", "post": post},
    )


@admin_required
def post_delete(request, pk):
    post = get_object_or_404(Post, pk=pk)
    if request.method == "POST":
        post.delete()
        messages.success(request, "Noticia eliminada")
    return redirect("feed")


@admin_required
def post_import(request):
    if request.method != "POST":
        return redirect("feed")

    form = ImportForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Selecciona un archivo válido (.xlsx o .csv)")
        return redirect("feed")

    archivo = form.cleaned_data["archivo"]
    name = archivo.name.lower()
    try:
        rows = _read_rows(archivo, name)
    except Exception:
        messages.error(
            request, "No se pudo leer el archivo. Verifica que sea .xlsx o .csv válido."
        )
        return redirect("feed")

    creadas = 0
    for row in rows:
        title = (row.get("titulo") or row.get("título") or row.get("title") or "").strip()
        if not title:
            continue
        Post.objects.create(
            title=title[:200],
            excerpt=(row.get("texto") or row.get("resumen") or row.get("excerpt") or "").strip()[:300],
            body=(row.get("contenido") or row.get("body") or "").strip(),
            category=_norm_category(row.get("categoria") or row.get("categoría") or row.get("category")),
            author=request.user,
        )
        creadas += 1

    if creadas:
        messages.success(request, f"Se importaron {creadas} noticia(s) desde el archivo")
    else:
        messages.error(
            request,
            "No se importó nada. Asegúrate de que la primera fila tenga la columna 'Titulo'.",
        )
    return redirect("feed")


@admin_required
def post_import_template(request):
    """Descarga un CSV de ejemplo (se abre en Excel) con las columnas correctas."""
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="plantilla_noticias.csv"'
    response.write("﻿")  # BOM para que Excel respete acentos
    writer = csv.writer(response)
    writer.writerow(["Titulo", "Texto", "Categoria"])
    writer.writerow([
        "Resultados del trimestre",
        "Tuvimos un trimestre récord. Gracias a todos los socios.",
        "Noticias",
    ])
    writer.writerow([
        "Convocatoria a Asamblea",
        "Se convoca a la asamblea ordinaria. Su asistencia es importante.",
        "Circular",
    ])
    return response


def _read_rows(archivo, name):
    """Devuelve una lista de dicts {columna_minuscula: valor} desde xlsx o csv."""
    rows = []
    if name.endswith(".csv"):
        data = archivo.read().decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(data))
        for r in reader:
            rows.append({(k or "").strip().lower(): (v or "") for k, v in r.items()})
    else:  # .xlsx
        from openpyxl import load_workbook

        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        header = []
        for i, raw in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c).strip().lower() if c is not None else "" for c in raw]
                continue
            row = {}
            for j, val in enumerate(raw):
                if j < len(header) and header[j]:
                    row[header[j]] = "" if val is None else str(val)
            rows.append(row)
        wb.close()
    return rows
